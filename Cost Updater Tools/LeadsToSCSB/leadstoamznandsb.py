import pandas as pd
import random
import string
from dotenv import load_dotenv
import os
import requests
import io
from io import StringIO, BytesIO
from openpyxl import load_workbook
import boto3
import json
import smtplib
import ssl
from email.message import EmailMessage

# Load environment variables
load_dotenv()

# Google Sheet URLs for Tevin and David
TEVIN_SHEET = os.getenv('TEVIN_SHEET')
DAVID_SHEET = os.getenv('DAVID_SHEET')  # New environment variable for David's sheet
OSCAR_SHEET = os.getenv('OSCAR_SHEET')

CONFIG_S3_BUCKET = os.getenv("CONFIG_S3_BUCKET")
LISTING_LOADER_KEY = "listingLoaderTemplate.xlsm"

# SB file keys and filenames for Tevin and David
TEVIN_SB_FILE_KEY = "tevin_sb.xlsx"
TEVIN_SB_UPDATED_FILE = "tevin_sb.xlsx"
DAVID_SB_FILE_KEY = "david_sb.xlsx"
DAVID_SB_UPDATED_FILE = "david_sb.xlsx"
OSCAR_SB_FILE_KEY = "oscar_sb.xlsx"
OSCAR_SB_UPDATED_FILE = "oscar_sb.xlsx"

# Email credentials and recipient addresses for Tevin and David
EMAIL_ADDRESS = os.getenv("EMAIL_ADDRESS")
EMAIL_PASSWORD = os.getenv("EMAIL_PASSWORD")
TEVIN_EMAIL = os.getenv("TEVIN_EMAIL")
DAVID_EMAIL = os.getenv("DAVID_EMAIL")
OSCAR_EMAIL = os.getenv("OSCAR_EMAIL")

if not TEVIN_SHEET:
    raise ValueError("TEVIN_SHEET environment variable is not set.")

def get_last_processed_date():
    """Retrieve the last processed date from the config file in S3."""
    s3_client = boto3.client('s3')
    config_key = "amznUploadConfig.json"
    try:
        response = s3_client.get_object(Bucket=CONFIG_S3_BUCKET, Key=config_key)
        config_data = json.loads(response['Body'].read().decode('utf-8'))
        return config_data.get("last_processed_date", "2000-01-01")
    except Exception as e:
        print(f"Error fetching last processed date: {e}")
        return "2000-01-01"

def update_last_processed_date(new_date):
    """Update the last processed date in the config file stored in S3."""
    s3_client = boto3.client('s3')
    config_key = "amznUploadConfig.json"
    new_config = json.dumps({"last_processed_date": new_date})
    try:
        s3_client.put_object(Bucket=CONFIG_S3_BUCKET, Key=config_key, Body=new_config)
        print(f"Updated last processed date to: {new_date} in S3.")
    except Exception as e:
        print(f"Error updating last processed date: {e}")

def send_email(attachments, recipient_email, potential_updates, new_products, actual_updates):
    """
    Sends email with multiple attachments and a report.
    attachments: list of tuples (BytesIO_object, filename)
    """
    msg = EmailMessage()
    msg['From'] = EMAIL_ADDRESS
    msg['To'] = recipient_email
    msg['Subject'] = "Amazon New Listings & COGS Report"
    
    # Create HTML content
    html_content = """<html>
    <body>
        <h2 style="color: #2c3e50;">Amazon New Listings & COGS Report</h2>
        <div style="margin-bottom: 30px;">"""

    if actual_updates:
        html_content += """
        <h3 style="color: #34495e;">Completed Cost Updates</h3>
        <table style="border-collapse: collapse; width: 100%; margin-bottom: 20px;">
            <tr style="background-color: #f8f9fa;">
                <th style="padding: 12px; border: 1px solid #ddd;">ASIN</th>
                <th style="padding: 12px; border: 1px solid #ddd;">SKU</th>
                <th style="padding: 12px; border: 1px solid #ddd;">Name</th>
                <th style="padding: 12px; border: 1px solid #ddd;">New Cost</th>
            </tr>"""
        for update in actual_updates:
            html_content += f"""
            <tr>
                <td style="padding: 12px; border: 1px solid #ddd;">{update['ASIN']}</td>
                <td style="padding: 12px; border: 1px solid #ddd;">{update['SKU']}</td>
                <td style="padding: 12px; border: 1px solid #ddd;">{update['Name']}</td>
                <td style="padding: 12px; border: 1px solid #ddd;">${update['new_cost']:.2f}</td>
            </tr>"""
        html_content += "</table>"
    
    if potential_updates:
        html_content += """
        <h3 style="color: #34495e;">Potential COGS Updates</h3>
        <table style="border-collapse: collapse; width: 100%; margin-bottom: 20px;">
            <tr style="background-color: #f8f9fa;">
                <th style="padding: 12px; border: 1px solid #ddd;">ASIN</th>
                <th style="padding: 12px; border: 1px solid #ddd;">SKU</th>
                <th style="padding: 12px; border: 1px solid #ddd;">Name</th>
                <th style="padding: 12px; border: 1px solid #ddd;">Old Cost</th>
                <th style="padding: 12px; border: 1px solid #ddd;">New Cost</th>
                <th style="padding: 12px; border: 1px solid #ddd;">Difference</th>
            </tr>"""
        for update in potential_updates:
            diff = update['new_cost'] - update['old_cost']
            diff_color = "#e74c3c" if diff > 0 else "#27ae60"  # Red for increase, Green for decrease
            html_content += f"""
            <tr>
                <td style="padding: 12px; border: 1px solid #ddd;">{update['ASIN']}</td>
                <td style="padding: 12px; border: 1px solid #ddd;">{update['SKU']}</td>
                <td style="padding: 12px; border: 1px solid #ddd;">{update['Name']}</td>
                <td style="padding: 12px; border: 1px solid #ddd;">${update['old_cost']:.2f}</td>
                <td style="padding: 12px; border: 1px solid #ddd;">${update['new_cost']:.2f}</td>
                <td style="padding: 12px; border: 1px solid #ddd; color: {diff_color};">
                    {diff:+.2f}
                </td>
            </tr>"""
        html_content += "</table>"
    
    if new_products:
        html_content += """
        <h3 style="color: #34495e;">New Products Added</h3>
        <table style="border-collapse: collapse; width: 100%;">
            <tr style="background-color: #f8f9fa;">
                <th style="padding: 12px; border: 1px solid #ddd;">ASIN</th>
                <th style="padding: 12px; border: 1px solid #ddd;">SKU</th>
                <th style="padding: 12px; border: 1px solid #ddd;">Name</th>
                <th style="padding: 12px; border: 1px solid #ddd;">Initial Cost</th>
            </tr>"""
        for product in new_products:
            html_content += f"""
            <tr>
                <td style="padding: 12px; border: 1px solid #ddd;">{product['ASIN']}</td>
                <td style="padding: 12px; border: 1px solid #ddd;">{product['SKU']}</td>
                <td style="padding: 12px; border: 1px solid #ddd;">{product['Name']}</td>
                <td style="padding: 12px; border: 1px solid #ddd;">${product['cost']:.2f}</td>
            </tr>"""
        html_content += "</table>"
    
    html_content += """
        </div>
        <p style="color: #7f8c8d;">
            Note: Potential COGS updates are suggestions only. No actual changes have been made to existing items.
            <br>Attached files contain new product listings and update prices of new listings within sellerboard.
        </p>
    </body>
    </html>"""
    
    msg.add_alternative(html_content, subtype='html')
    
    # Attach each file from the attachments list
    for attachment_data, attachment_filename in attachments:
        attachment_data.seek(0)
        try:
            msg.add_attachment(
                attachment_data.read(),
                filename=attachment_filename,
                maintype="application",
                subtype="octet-stream"
            )
        except Exception as e:
            print(f"Failed to add attachment {attachment_filename}: {e}")
    
    try:
        context = ssl.create_default_context()
        with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
            server.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
            server.send_message(msg)
        print(f"Email sent successfully to {recipient_email} with all attachments.")
    except Exception as e:
        print(f"Failed to send email to {recipient_email}: {e}")

def fetch_google_sheet(url):
    """Fetches the Google Sheet CSV data and returns a pandas DataFrame."""
    response = requests.get(url)
    response.raise_for_status()
    csv_data = StringIO(response.text)
    return pd.read_csv(csv_data, dtype=str)

def generate_sku():
    """Generates a random SKU in the format: 4 letters - (3 numbers + 3 letters)."""
    letters = ''.join(random.choices(string.ascii_uppercase, k=4))
    mixed_part = ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))
    return f"{letters}-{mixed_part}"

def fetch_s3_file(bucket, key):
    """Fetches a file from S3 and returns it as a BytesIO object."""
    s3_client = boto3.client('s3')
    response = s3_client.get_object(Bucket=bucket, Key=key)
    return BytesIO(response['Body'].read())

def process_sheet(sheet_url, sb_file_key, sb_updated_file, ws, headers, col_indices, last_processed_date):
    """
    Process a Google Sheet (either Tevin or David) and update its corresponding SB DataFrame.
    Returns the updated DataFrame, SB DataFrame, and lists of updates.
    """
    df = fetch_google_sheet(sheet_url)
    df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
    # Filter rows that are after the last processed date
    df = df[df['Date'] >= pd.to_datetime(last_processed_date)]
    df['Sale Price'] = df['Sale Price'].astype(str).str.replace('$', '').str.strip()
    
    sb_data = fetch_s3_file(CONFIG_S3_BUCKET, sb_file_key)
    sb_df = pd.read_excel(sb_data)
    sb_df.columns = sb_df.columns.str.strip()
    sb_df['ASIN'] = sb_df['ASIN'].astype(str).str.strip()
    sb_df['SKU'] = sb_df['SKU'].astype(str).str.strip()
    df['ASIN'] = df['ASIN'].astype(str).str.strip()
    df['COGS'] = pd.to_numeric(df['COGS'].replace(r'[\$,]', '', regex=True), errors='coerce')
    
    potential_updates = []
    new_products = []
    actual_updates = []
    
    for _, row in df.iterrows():
        if row['Sale Price'] == "Replen":
            continue

        asin = row['ASIN']
        existing_entry = sb_df[sb_df['ASIN'] == asin]
        name = row['Name']
        try:
            new_cost = float(row['COGS'])
        except (ValueError, TypeError):
            continue

        if not existing_entry.empty:
            old_cost = existing_entry.iloc[0]['Cost']
            sku = existing_entry.iloc[0]['SKU']
            if pd.isna(old_cost) or old_cost == '':
                sb_df.loc[sb_df['ASIN'] == asin, 'Cost'] = new_cost
                actual_updates.append({
                    'ASIN': asin,
                    'SKU': sku,
                    'Name': name,
                    'new_cost': new_cost
                })
            else:
                old_cost_val = float(old_cost)
                # Only add to potential_updates if there's a difference in cost
                if new_cost != old_cost_val:
                    potential_updates.append({
                        'ASIN': asin,
                        'SKU': sku,
                        'Name': name,
                        'old_cost': old_cost_val,
                        'new_cost': new_cost
                    })
        else:
            sku = generate_sku()
            new_products.append({
                'ASIN': asin,
                'SKU': sku,
                'Name': name,
                'cost': new_cost
            })
            new_sb_row = {
                'ASIN': asin,
                'SKU': sku,
                'Title': name,
                'Labels': '#FBA',
                'Cost': new_cost,
                'VAT_CATEGORY': 'A_GEN_STANDARD',
                'Hide': 'NO'
            }
            sb_df = pd.concat([sb_df, pd.DataFrame([new_sb_row])], ignore_index=True)

            # Add to Listing Loader (new products only)
            new_row = [None] * len(headers)
            new_row[col_indices['Your Search Term'] - 1] = asin
            new_row[col_indices['Recommended Action'] - 1] = 'Ready To list > Enter required details.'
            new_row[col_indices["Amazon's Title"] - 1] = row['Name']
            new_row[col_indices['Record Action'] - 1] = 'Add Product'
            new_row[col_indices['Seller SKU'] - 1] = sku
            new_row[col_indices['Merchant Suggested ASIN'] - 1] = asin
            new_row[col_indices['Offering Condition Type'] - 1] = 'New'
            new_row[col_indices['Fulfillment Channel Code (US)'] - 1] = 'AMAZON_NA'
            new_row[col_indices['Your Price USD (Sell on Amazon, US)'] - 1] = f"{float(row['Sale Price']) * 1.15:.2f}"
            ws.append(new_row)

    
    return df, sb_df, potential_updates, new_products, actual_updates

def lambda_handler(event, context):
    """AWS Lambda entry point."""
    try:
        last_processed_date = get_last_processed_date()
        new_date_list = []
        
        # Load Listing Loader workbook from S3 (common for both sheets)
        listing_loader_data = fetch_s3_file(CONFIG_S3_BUCKET, LISTING_LOADER_KEY)
        wb = load_workbook(filename=listing_loader_data, keep_vba=True)
        ws = wb["Template"]
        headers = [cell.value for cell in ws[4]]
        
        def get_column_index(col_name):
            """Finds the column index of a given column name."""
            return headers.index(col_name) + 1 if col_name in headers else None
        
        col_indices = {col: get_column_index(col) for col in [
            'Your Search Term', "Amazon's Title", 'Record Action', 'Seller SKU', 
            'Merchant Suggested ASIN', 'Offering Condition Type', 'Fulfillment Channel Code (US)',
            'Your Price USD (Sell on Amazon, US)', 'Recommended Action'
        ]}
        
        # Initialize lists to collect updates for both sheets
        potential_updates_all = []
        new_products_all = []
        actual_updates_all = []
        
        # Process Tevin Sheet
        if TEVIN_SHEET:
            tevin_df, tevin_sb_df, potential_updates_tevin, new_products_tevin, actual_updates_tevin = process_sheet(
                TEVIN_SHEET, TEVIN_SB_FILE_KEY, TEVIN_SB_UPDATED_FILE, ws, headers, col_indices, last_processed_date
            )
            potential_updates_all.extend(potential_updates_tevin)
            new_products_all.extend(new_products_tevin)
            actual_updates_all.extend(actual_updates_tevin)
            if not tevin_df.empty:
                new_date_list.append(tevin_df["Date"].max())
        else:
            tevin_sb_df = None
        
        # Process David Sheet if provided
        if DAVID_SHEET:
            david_df, david_sb_df, potential_updates_david, new_products_david, actual_updates_david = process_sheet(
                DAVID_SHEET, DAVID_SB_FILE_KEY, DAVID_SB_UPDATED_FILE, ws, headers, col_indices, last_processed_date
            )
            potential_updates_all.extend(potential_updates_david)
            new_products_all.extend(new_products_david)
            actual_updates_all.extend(actual_updates_david)
            if not david_df.empty:
                new_date_list.append(david_df["Date"].max())
        else:
            david_sb_df = None

        if OSCAR_SHEET:
            oscar_df, oscar_sb_df, potential_updates_oscar, new_products_oscar, actual_updates_oscar = process_sheet(
                OSCAR_SHEET, OSCAR_SB_FILE_KEY, OSCAR_SB_UPDATED_FILE, ws, headers, col_indices, last_processed_date
            )
            potential_updates_all.extend(potential_updates_oscar)
            new_products_all.extend(new_products_oscar)
            actual_updates_all.extend(actual_updates_oscar)
            if not oscar_df.empty:
                new_date_list.append(oscar_df["Date"].max())
        else:
            oscar_sb_df = None
        
        # Save updated Listing Loader workbook into a BytesIO buffer
        listing_loader_output_buffer = io.BytesIO()
        wb.save(listing_loader_output_buffer)
        listing_loader_output_buffer.seek(0)
        # Create a separate BytesIO for each email from the same data
        listing_loader_bytes = listing_loader_output_buffer.getvalue()
        
        # Prepare updated SB file buffers and attachments for Tevin
        attachments_tevin = []
        if tevin_sb_df is not None:
            tevin_sb_buffer = io.BytesIO()
            tevin_sb_df.to_excel(tevin_sb_buffer, index=False, engine='openpyxl')
            tevin_sb_buffer.seek(0)
            attachments_tevin = [
                (BytesIO(listing_loader_bytes), "listingLoaderUpdated.xlsm"),
                (tevin_sb_buffer, TEVIN_SB_UPDATED_FILE)
            ]
        
        # Prepare updated SB file buffers and attachments for David
        attachments_david = []
        if david_sb_df is not None:
            david_sb_buffer = io.BytesIO()
            david_sb_df.to_excel(david_sb_buffer, index=False, engine='openpyxl')
            david_sb_buffer.seek(0)
            attachments_david = [
                (BytesIO(listing_loader_bytes), "listingLoaderUpdated.xlsm"),
                (david_sb_buffer, DAVID_SB_UPDATED_FILE)
            ]

        attachments_oscar = []
        if oscar_sb_df is not None:
            oscar_sb_buffer = io.BytesIO()
            oscar_sb_df.to_excel(oscar_sb_buffer, index=False, engine='openpyxl')
            oscar_sb_buffer.seek(0)
            attachments_oscar = [
                (BytesIO(listing_loader_bytes), "listingLoaderUpdated.xlsm"),
                (oscar_sb_buffer, OSCAR_SB_UPDATED_FILE)
            ]
        
        # Upload updated SB files to S3
        s3_client = boto3.client('s3')
        try:
            if tevin_sb_df is not None:
                s3_client.put_object(Bucket=CONFIG_S3_BUCKET, Key=TEVIN_SB_FILE_KEY, Body=tevin_sb_buffer.getvalue())
                print("Successfully uploaded updated Tevin SB file to S3")
            if david_sb_df is not None:
                s3_client.put_object(Bucket=CONFIG_S3_BUCKET, Key=DAVID_SB_FILE_KEY, Body=david_sb_buffer.getvalue())
                print("Successfully uploaded updated David SB file to S3")
            if oscar_sb_df is not None:
                s3_client.put_object(Bucket=CONFIG_S3_BUCKET, Key=OSCAR_SB_FILE_KEY, Body=oscar_sb_buffer.getvalue())
                print("Successfully uploaded updated Oscar SB file to S3")
        except Exception as e:
            print(f"Error uploading SB file to S3: {e}")
            raise
        
        # Send email to Tevin if available
        if attachments_tevin and TEVIN_EMAIL:
            send_email(attachments_tevin, TEVIN_EMAIL, potential_updates_all, new_products_all, actual_updates_all)
        
        # Send email to David if available
        if attachments_david and DAVID_EMAIL:
            send_email(attachments_david, DAVID_EMAIL, potential_updates_all, new_products_all, actual_updates_all)
        
        if attachments_oscar and OSCAR_EMAIL:
            send_email(attachments_oscar, OSCAR_EMAIL, potential_updates_all, new_products_all, actual_updates_all)
        

        # Update last processed date using the maximum date found in both sheets
        if new_date_list:
            new_last_processed_date = str(pd.to_datetime(max(new_date_list)).date())
            update_last_processed_date(new_last_processed_date)
        
        return {
            'statusCode': 200,
            'body': json.dumps('Process completed successfully!')
        }
        
    except Exception as e:
        print(f"Error: {e}")
        return {
            'statusCode': 500,
            'body': json.dumps(f"Error: {str(e)}")
        }
