import pandas as pd
import random
import string
from dotenv import load_dotenv
import os
import requests
import io
from io import StringIO, BytesIO
from openpyxl import load_workbook
import boto3
import json
import smtplib
import ssl
from email.message import EmailMessage

# Load environment variables
load_dotenv()

# Common configuration variables
CONFIG_S3_BUCKET = os.getenv("CONFIG_S3_BUCKET")
LISTING_LOADER_KEY = "listingLoaderTemplate.xlsm"
EMAIL_ADDRESS = os.getenv("EMAIL_ADDRESS")
EMAIL_PASSWORD = os.getenv("EMAIL_PASSWORD")

# User configuration: add new users by simply appending a dictionary to this list.
users_config = [
    {
        "name": "Tevin",
        "sheet_url": os.getenv("TEVIN_SHEET"),
        "sb_file_key": "tevin_sb.xlsx",
        "sb_updated_file": "tevin_sb.xlsx",
        "email": os.getenv("TEVIN_EMAIL")
    },
    {
        "name": "David",
        "sheet_url": os.getenv("DAVID_SHEET"),
        "sb_file_key": "david_sb.xlsx",
        "sb_updated_file": "david_sb.xlsx",
        "email": os.getenv("DAVID_EMAIL")
    },
    {
        "name": "Oscar",
        "sheet_url": os.getenv("OSCAR_SHEET"),
        "sb_file_key": "oscar_sb.xlsx",
        "sb_updated_file": "oscar_sb.xlsx",
        "email": os.getenv("OSCAR_EMAIL")
    },
    # Add more users here as needed.
]

if not any(user["sheet_url"] for user in users_config):
    raise ValueError("At least one user sheet URL must be set.")

def get_last_processed_date():
    s3_client = boto3.client('s3')
    config_key = "amznUploadConfig.json"
    try:
        response = s3_client.get_object(Bucket=CONFIG_S3_BUCKET, Key=config_key)
        config_data = json.loads(response['Body'].read().decode('utf-8'))
        return config_data.get("last_processed_date", "2000-01-01")
    except Exception as e:
        print(f"Error fetching last processed date: {e}")
        return "2000-01-01"

def update_last_processed_date(new_date):
    s3_client = boto3.client('s3')
    config_key = "amznUploadConfig.json"
    new_config = json.dumps({"last_processed_date": new_date})
    try:
        s3_client.put_object(Bucket=CONFIG_S3_BUCKET, Key=config_key, Body=new_config)
        print(f"Updated last processed date to: {new_date} in S3.")
    except Exception as e:
        print(f"Error updating last processed date: {e}")

def send_email(attachments, recipient_email, potential_updates, new_products, actual_updates):
    """
    Sends email with multiple attachments and a report.
    attachments: list of tuples (BytesIO_object, filename)
    """
    msg = EmailMessage()
    msg['From'] = EMAIL_ADDRESS
    msg['To'] = recipient_email
    msg['Subject'] = "Amazon New Listings & COGS Report"
    
    # Build HTML email content
    html_content = """<html>
    <body>
        <h2 style="color: #2c3e50;">Amazon New Listings & COGS Report</h2>
        <div style="margin-bottom: 30px;">"""
    
    if actual_updates:
        html_content += """
        <h3 style="color: #34495e;">Completed Cost Updates</h3>
        <table style="border-collapse: collapse; width: 100%; margin-bottom: 20px;">
            <tr style="background-color: #f8f9fa;">
                <th style="padding: 12px; border: 1px solid #ddd;">ASIN</th>
                <th style="padding: 12px; border: 1px solid #ddd;">SKU</th>
                <th style="padding: 12px; border: 1px solid #ddd;">Name</th>
                <th style="padding: 12px; border: 1px solid #ddd;">New Cost</th>
            </tr>"""
        for update in actual_updates:
            html_content += f"""
            <tr>
                <td style="padding: 12px; border: 1px solid #ddd;">{update['ASIN']}</td>
                <td style="padding: 12px; border: 1px solid #ddd;">{update['SKU']}</td>
                <td style="padding: 12px; border: 1px solid #ddd;">{update['Name']}</td>
                <td style="padding: 12px; border: 1px solid #ddd;">${update['new_cost']:.2f}</td>
            </tr>"""
        html_content += "</table>"
    
    if potential_updates:
        html_content += """
        <h3 style="color: #34495e;">Potential COGS Updates</h3>
        <table style="border-collapse: collapse; width: 100%; margin-bottom: 20px;">
            <tr style="background-color: #f8f9fa;">
                <th style="padding: 12px; border: 1px solid #ddd;">ASIN</th>
                <th style="padding: 12px; border: 1px solid #ddd;">SKU</th>
                <th style="padding: 12px; border: 1px solid #ddd;">Name</th>
                <th style="padding: 12px; border: 1px solid #ddd;">Old Cost</th>
                <th style="padding: 12px; border: 1px solid #ddd;">New Cost</th>
                <th style="padding: 12px; border: 1px solid #ddd;">Difference</th>
            </tr>"""
        for update in potential_updates:
            diff = update['new_cost'] - update['old_cost']
            diff_color = "#e74c3c" if diff > 0 else "#27ae60"
            html_content += f"""
            <tr>
                <td style="padding: 12px; border: 1px solid #ddd;">{update['ASIN']}</td>
                <td style="padding: 12px; border: 1px solid #ddd;">{update['SKU']}</td>
                <td style="padding: 12px; border: 1px solid #ddd;">{update['Name']}</td>
                <td style="padding: 12px; border: 1px solid #ddd;">${update['old_cost']:.2f}</td>
                <td style="padding: 12px; border: 1px solid #ddd;">${update['new_cost']:.2f}</td>
                <td style="padding: 12px; border: 1px solid #ddd; color: {diff_color};">
                    {diff:+.2f}
                </td>
            </tr>"""
        html_content += "</table>"
    
    if new_products:
        html_content += """
        <h3 style="color: #34495e;">New Products Added</h3>
        <table style="border-collapse: collapse; width: 100%;">
            <tr style="background-color: #f8f9fa;">
                <th style="padding: 12px; border: 1px solid #ddd;">ASIN</th>
                <th style="padding: 12px; border: 1px solid #ddd;">SKU</th>
                <th style="padding: 12px; border: 1px solid #ddd;">Name</th>
                <th style="padding: 12px; border: 1px solid #ddd;">Initial Cost</th>
            </tr>"""
        for product in new_products:
            html_content += f"""
            <tr>
                <td style="padding: 12px; border: 1px solid #ddd;">{product['ASIN']}</td>
                <td style="padding: 12px; border: 1px solid #ddd;">{product['SKU']}</td>
                <td style="padding: 12px; border: 1px solid #ddd;">{product['Name']}</td>
                <td style="padding: 12px; border: 1px solid #ddd;">${product['cost']:.2f}</td>
            </tr>"""
        html_content += "</table>"
    
    html_content += """
        </div>
        <p style="color: #7f8c8d;">
            Note: Potential COGS updates are suggestions only. No actual changes have been made to existing items.
            <br>Attached files contain new product listings and updated prices within Sellerboard.
        </p>
    </body>
    </html>"""
    
    msg.add_alternative(html_content, subtype='html')
    
    # Attach each file from the attachments list
    for attachment_data, attachment_filename in attachments:
        attachment_data.seek(0)
        try:
            msg.add_attachment(
                attachment_data.read(),
                filename=attachment_filename,
                maintype="application",
                subtype="octet-stream"
            )
        except Exception as e:
            print(f"Failed to add attachment {attachment_filename}: {e}")
    
    try:
        context = ssl.create_default_context()
        with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
            server.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
            server.send_message(msg)
        print(f"Email sent successfully to {recipient_email} with all attachments.")
    except Exception as e:
        print(f"Failed to send email to {recipient_email}: {e}")

def fetch_google_sheet(url):
    """Fetches the Google Sheet CSV data and returns a pandas DataFrame."""
    response = requests.get(url)
    response.raise_for_status()
    csv_data = StringIO(response.text)
    return pd.read_csv(csv_data, dtype=str)

def generate_sku():
    """Generates a random SKU in the format: 4 letters - (6 characters mix)."""
    letters = ''.join(random.choices(string.ascii_uppercase, k=4))
    mixed_part = ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))
    return f"{letters}-{mixed_part}"

def fetch_s3_file(bucket, key):
    """Fetches a file from S3 and returns it as a BytesIO object."""
    s3_client = boto3.client('s3')
    response = s3_client.get_object(Bucket=bucket, Key=key)
    return BytesIO(response['Body'].read())

def process_sheet(sheet_url, sb_file_key, sb_updated_file, ws, headers, col_indices, last_processed_date):
    """
    Process a Google Sheet and update its corresponding Sellerboard DataFrame.
    Returns the sheet DataFrame, the updated SB DataFrame, and lists of updates.
    """
    df = fetch_google_sheet(sheet_url)
    df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
    # Filter rows that are after the last processed date
    df = df[df['Date'] >= pd.to_datetime(last_processed_date)]
    df['Sale Price'] = df['Sale Price'].astype(str).str.replace('$', '').str.strip()
    
    sb_data = fetch_s3_file(CONFIG_S3_BUCKET, sb_file_key)
    sb_df = pd.read_excel(sb_data)
    sb_df.columns = sb_df.columns.str.strip()
    sb_df['ASIN'] = sb_df['ASIN'].astype(str).str.strip()
    sb_df['SKU'] = sb_df['SKU'].astype(str).str.strip()
    df['ASIN'] = df['ASIN'].astype(str).str.strip()
    df['COGS'] = pd.to_numeric(df['COGS'].replace(r'[\$,]', '', regex=True), errors='coerce')
    
    potential_updates = []
    new_products = []
    actual_updates = []
    
    for _, row in df.iterrows():
        if row['Sale Price'] == "Replen":
            continue

        asin = row['ASIN']
        existing_entry = sb_df[sb_df['ASIN'] == asin]
        name = row['Name']
        try:
            new_cost = float(row['COGS'])
        except (ValueError, TypeError):
            continue

        if not existing_entry.empty:
            old_cost = existing_entry.iloc[0]['Cost']
            sku = existing_entry.iloc[0]['SKU']
            if pd.isna(old_cost) or old_cost == '':
                sb_df.loc[sb_df['ASIN'] == asin, 'Cost'] = new_cost
                actual_updates.append({
                    'ASIN': asin,
                    'SKU': sku,
                    'Name': name,
                    'new_cost': new_cost
                })
            else:
                old_cost_val = float(old_cost)
                if new_cost != old_cost_val:
                    potential_updates.append({
                        'ASIN': asin,
                        'SKU': sku,
                        'Name': name,
                        'old_cost': old_cost_val,
                        'new_cost': new_cost
                    })
        else:
            sku = generate_sku()
            new_products.append({
                'ASIN': asin,
                'SKU': sku,
                'Name': name,
                'cost': new_cost
            })
            new_sb_row = {
                'ASIN': asin,
                'SKU': sku,
                'Title': name,
                'Labels': '#FBA',
                'Cost': new_cost,
                'VAT_CATEGORY': 'A_GEN_STANDARD',
                'Hide': 'NO'
            }
            sb_df = pd.concat([sb_df, pd.DataFrame([new_sb_row])], ignore_index=True)
            new_row = [None] * len(headers)
            new_row[col_indices['Your Search Term'] - 1] = asin
            new_row[col_indices['Recommended Action'] - 1] = 'Ready To list > Enter required details.'
            new_row[col_indices["Amazon's Title"] - 1] = row['Name']
            new_row[col_indices['Record Action'] - 1] = 'Add Product'
            new_row[col_indices['Seller SKU'] - 1] = sku
            new_row[col_indices['Merchant Suggested ASIN'] - 1] = asin
            new_row[col_indices['Offering Condition Type'] - 1] = 'New'
            new_row[col_indices['Fulfillment Channel Code (US)'] - 1] = 'AMAZON_NA'
            new_row[col_indices['Your Price USD (Sell on Amazon, US)'] - 1] = f"{float(row['Sale Price']) * 1.15:.2f}"
            ws.append(new_row)
    
    return df, sb_df, potential_updates, new_products, actual_updates

def lambda_handler(event, context):
    """AWS Lambda entry point."""
    try:
        last_processed_date = get_last_processed_date()
        new_date_list = []
        s3_client = boto3.client('s3')
        
        # Process each user separately so each gets a unique Listing Loader
        for user in users_config:
            if not user["sheet_url"]:
                continue
            
            # Load a fresh copy of the Listing Loader workbook for this user
            listing_loader_data = fetch_s3_file(CONFIG_S3_BUCKET, LISTING_LOADER_KEY)
            wb = load_workbook(filename=listing_loader_data, keep_vba=True)
            ws = wb["Template"]
            headers = [cell.value for cell in ws[4]]
            
            def get_column_index(col_name):
                return headers.index(col_name) + 1 if col_name in headers else None
            
            col_indices = {col: get_column_index(col) for col in [
                'Your Search Term', "Amazon's Title", 'Record Action', 'Seller SKU', 
                'Merchant Suggested ASIN', 'Offering Condition Type', 'Fulfillment Channel Code (US)',
                'Your Price USD (Sell on Amazon, US)', 'Recommended Action'
            ]}
            
            # Process the user's Google Sheet
            df, sb_df, potential_updates, new_products, actual_updates = process_sheet(
                user["sheet_url"],
                user["sb_file_key"],
                user["sb_updated_file"],
                ws,
                headers,
                col_indices,
                last_processed_date
            )
            if not df.empty:
                new_date_list.append(df["Date"].max())
            
            # Save this user's updated Listing Loader workbook to a buffer
            listing_loader_output_buffer = io.BytesIO()
            wb.save(listing_loader_output_buffer)
            listing_loader_output_buffer.seek(0)
            listing_loader_bytes = listing_loader_output_buffer.getvalue()
            
            # Upload updated Sellerboard file for this user to S3
            sb_buffer = io.BytesIO()
            sb_df.to_excel(sb_buffer, index=False, engine='openpyxl')
            sb_buffer.seek(0)
            s3_client.put_object(Bucket=CONFIG_S3_BUCKET, Key=user["sb_file_key"], Body=sb_buffer.getvalue())
            print(f"Successfully uploaded updated {user['name']} SB file to S3")
            
            # Prepare attachments for this user: each gets their own Listing Loader workbook
            attachments = [
                (BytesIO(listing_loader_bytes), "listingLoaderUpdated.xlsm"),
                (sb_buffer, user["sb_updated_file"])
            ]
            
            if user["email"]:
                send_email(
                    attachments,
                    user["email"],
                    potential_updates,
                    new_products,
                    actual_updates
                )
        
        # Update the last processed date using the maximum date from all users
        if new_date_list:
            new_last_processed_date = str(pd.to_datetime(max(new_date_list)).date())
            update_last_processed_date(new_last_processed_date)
        
        return {
            'statusCode': 200,
            'body': json.dumps('Process completed successfully!')
        }
    
    except Exception as e:
        print(f"Error: {e}")
        return {
            'statusCode': 500,
            'body': json.dumps(f"Error: {str(e)}")
        }