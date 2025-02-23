import pandas as pd
import random
import string
from dotenv import load_dotenv
import os
import requests
import io
from io import StringIO, BytesIO
from openpyxl import load_workbook
import boto3
import json
import smtplib
import ssl
from email.message import EmailMessage

# Load environment variables
load_dotenv()

# Fetch the Google Sheet URL from environment variables
TEVIN_SHEET = os.getenv('TEVIN_SHEET')
CONFIG_S3_BUCKET = os.getenv("CONFIG_S3_BUCKET")
LISTING_LOADER_KEY = "listingLoaderTemplate.xlsm"
SB_FILE_KEY = "sb.xlsx"
SB_UPDATED_FILE = "sb.xlsx"

EMAIL_ADDRESS = os.getenv("EMAIL_ADDRESS")
EMAIL_PASSWORD = os.getenv("EMAIL_PASSWORD")

if not TEVIN_SHEET:
    raise ValueError("TEVIN_SHEET environment variable is not set.")

def get_last_processed_date():
    """Retrieve the last processed date from the config file in S3."""
    s3_client = boto3.client('s3')
    config_key = "amznUploadConfig.json"
    try:
        response = s3_client.get_object(Bucket=CONFIG_S3_BUCKET, Key=config_key)
        config_data = json.loads(response['Body'].read().decode('utf-8'))
        return config_data.get("last_processed_date", "2000-01-01")
    except Exception as e:
        print(f"Error fetching last processed date: {e}")
        return "2000-01-01"

def update_last_processed_date(new_date):
    """Update the last processed date in the config file stored in S3."""
    s3_client = boto3.client('s3')
    config_key = "amznUploadConfig.json"
    new_config = json.dumps({"last_processed_date": new_date})
    try:
        s3_client.put_object(Bucket=CONFIG_S3_BUCKET, Key=config_key, Body=new_config)
        print(f"Updated last processed date to: {new_date} in S3.")
    except Exception as e:
        print(f"Error updating last processed date: {e}")

def send_email(attachment_data1, attachment_filename1, 
              attachment_data2, attachment_filename2, 
              recipient_email, potential_updates, new_products, actual_updates):
    """Sends email with attachments and potential updates report."""
    msg = EmailMessage()
    msg['From'] = EMAIL_ADDRESS
    msg['To'] = recipient_email
    msg['Subject'] = "Amazon New Listings & COGS Report"
    
    # Create HTML content
    html_content = """<html>
    <body>
        <h2 style="color: #2c3e50;">Amazon New Listings & COGS Report</h2>
        <div style="margin-bottom: 30px;">"""

    if actual_updates:
        html_content += """
        <h3 style="color: #34495e;">Completed Cost Updates</h3>
        <table style="border-collapse: collapse; width: 100%; margin-bottom: 20px;">
            <tr style="background-color: #f8f9fa;">
                <th style="padding: 12px; border: 1px solid #ddd;">ASIN</th>
                <th style="padding: 12px; border: 1px solid #ddd;">SKU</th>
                <th style="padding: 12px; border: 1px solid #ddd;">Name</th>
                <th style="padding: 12px; border: 1px solid #ddd;">New Cost</th>
            </tr>"""
        for update in actual_updates:
            html_content += f"""
            <tr>
                <td style="padding: 12px; border: 1px solid #ddd;">{update['ASIN']}</td>
                <td style="padding: 12px; border: 1px solid #ddd;">{update['SKU']}</td>
                <td style="padding: 12px; border: 1px solid #ddd;">{update['Name']}</td>
                <td style="padding: 12px; border: 1px solid #ddd;">${update['new_cost']:.2f}</td>
            </tr>"""
        html_content += "</table>"
    
    if potential_updates:
        html_content += """
        <h3 style="color: #34495e;">Potential COGS Updates</h3>
        <table style="border-collapse: collapse; width: 100%; margin-bottom: 20px;">
            <tr style="background-color: #f8f9fa;">
                <th style="padding: 12px; border: 1px solid #ddd;">ASIN</th>
                <th style="padding: 12px; border: 1px solid #ddd;">SKU</th>
                <th style="padding: 12px; border: 1px solid #ddd;">Name</th>
                <th style="padding: 12px; border: 1px solid #ddd;">Old Cost</th>
                <th style="padding: 12px; border: 1px solid #ddd;">New Cost</th>
                <th style="padding: 12px; border: 1px solid #ddd;">Difference</th>
            </tr>"""
        for update in potential_updates:
            diff = update['new_cost'] - update['old_cost']
            diff_color = "#e74c3c" if diff > 0 else "#27ae60"  # Red for increase, Green for decrease
            html_content += f"""
            <tr>
                <td style="padding: 12px; border: 1px solid #ddd;">{update['ASIN']}</td>
                <td style="padding: 12px; border: 1px solid #ddd;">{update['SKU']}</td>
                <td style="padding: 12px; border: 1px solid #ddd;">{update['Name']}</td>
                <td style="padding: 12px; border: 1px solid #ddd;">${update['old_cost']:.2f}</td>
                <td style="padding: 12px; border: 1px solid #ddd;">${update['new_cost']:.2f}</td>
                <td style="padding: 12px; border: 1px solid #ddd; color: {diff_color};">
                    {diff:+.2f}
                </td>
            </tr>"""
        html_content += "</table>"
    
    if new_products:
        html_content += """
        <h3 style="color: #34495e;">New Products Added</h3>
        <table style="border-collapse: collapse; width: 100%;">
            <tr style="background-color: #f8f9fa;">
                <th style="padding: 12px; border: 1px solid #ddd;">ASIN</th>
                <th style="padding: 12px; border: 1px solid #ddd;">SKU</th>
                <th style="padding: 12px; border: 1px solid #ddd;">Name</th>
                <th style="padding: 12px; border: 1px solid #ddd;">Initial Cost</th>
            </tr>"""
        for product in new_products:
            html_content += f"""
            <tr>
                <td style="padding: 12px; border: 1px solid #ddd;">{product['ASIN']}</td>
                <td style="padding: 12px; border: 1px solid #ddd;">{product['SKU']}</td>
                <td style="padding: 12px; border: 1px solid #ddd;">{product['Name']}</td>
                <td style="padding: 12px; border: 1px solid #ddd;">${product['cost']:.2f}</td>
            </tr>"""
        html_content += "</table>"
    
    html_content += """
        </div>
        <p style="color: #7f8c8d;">
            Note: Potential COGS updates are suggestions only. No actual changes have been made to existing items.
            <br>Attached files contain new product listings and update prices of new listings within sellerboard.
        </p>
    </body>
    </html>"""
    
    msg.add_alternative(html_content, subtype='html')

    try:
        # Attach SB file
        attachment_data1.seek(0)  # Make sure the file pointer is at the start
        msg.add_attachment(
            attachment_data1.read(),  # Read the file content as binary
            filename=attachment_filename1,  # Filename
            maintype="application",  # MIME type
            subtype="octet-stream"  # Generic file type
        )
        
        # Attach Listing Loader file
        attachment_data2.seek(0)  # Make sure the file pointer is at the start
        msg.add_attachment(
            attachment_data2.read(),  # Read the file content as binary
            filename=attachment_filename2,  # Filename
            maintype="application",  # MIME type
            subtype="octet-stream"  # Generic file type
        )
    except Exception as e:
        print(f"Failed to add attachments: {e}")
        return

    try:
        context = ssl.create_default_context()
        with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
            server.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
            server.send_message(msg)
        print("Email sent successfully with both attachments.")
    except Exception as e:
        print(f"Failed to send email: {e}")

def fetch_google_sheet(url):
    """Fetches the Google Sheet CSV data and returns a pandas DataFrame."""
    response = requests.get(url)
    response.raise_for_status()
    csv_data = StringIO(response.text)
    return pd.read_csv(csv_data, dtype=str)

def generate_sku():
    """Generates a random SKU in the format: 4 letters - (3 numbers + 3 letters)."""
    letters = ''.join(random.choices(string.ascii_uppercase, k=4))
    mixed_part = ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))
    return f"{letters}-{mixed_part}"

def fetch_s3_file(bucket, key):
    """Fetches a file from S3 and returns it as a BytesIO object."""
    s3_client = boto3.client('s3')
    response = s3_client.get_object(Bucket=bucket, Key=key)
    return BytesIO(response['Body'].read())

def lambda_handler(event, context):
    """AWS Lambda entry point."""
    try:
        # Get the last processed date
        last_processed_date = get_last_processed_date()

        # Load leads sheet
        leads_df = fetch_google_sheet(TEVIN_SHEET)

        # Convert the 'Date' column to datetime (ensure you have a Date column in your leads sheet)
        leads_df['Date'] = pd.to_datetime(leads_df['Date'], errors='coerce')

        # Filter rows that are after the last processed date
        leads_df = leads_df[leads_df['Date'] >= pd.to_datetime(last_processed_date)]

        # Continue with processing only filtered data
        leads_df['Sale Price'] = leads_df['Sale Price'].astype(str).str.replace('$', '').str.strip()

        # Load ListingLoader file from S3
        listing_loader_data = fetch_s3_file(CONFIG_S3_BUCKET, LISTING_LOADER_KEY)
        wb = load_workbook(filename=listing_loader_data, keep_vba=True)
        ws = wb["Template"]
        headers = [cell.value for cell in ws[4]]

        def get_column_index(col_name):
            """Finds the column index of a given column name."""
            return headers.index(col_name) + 1 if col_name in headers else None

        col_indices = {col: get_column_index(col) for col in [
            'Your Search Term', "Amazon's Title", 'Record Action', 'Seller SKU', 
            'Merchant Suggested ASIN', 'Offering Condition Type', 'Fulfillment Channel Code (US)',
            'Your Price USD (Sell on Amazon, US)', 'Recommended Action'
        ]}

        # Load sb.xlsx from S3
        sb_data = fetch_s3_file(CONFIG_S3_BUCKET, SB_FILE_KEY)
        sb_df = pd.read_excel(sb_data)

        sb_df.columns = sb_df.columns.str.strip()
        sb_df['ASIN'] = sb_df['ASIN'].astype(str).str.strip()
        sb_df['SKU'] = sb_df['SKU'].astype(str).str.strip()
        leads_df['ASIN'] = leads_df['ASIN'].astype(str).str.strip()
        leads_df['COGS'] = pd.to_numeric(leads_df['COGS'].replace(r'[\$,]', '', regex=True), errors='coerce')

        potential_updates = []
        new_products = []
        actual_updates = []

        # Process leads data
        for _, row in leads_df.iterrows():
            if row['Sale Price'] == "Replen":
                continue

            asin = row['ASIN']
            existing_entry = sb_df[sb_df['ASIN'] == asin]
            name = row['Name']
            
            try:
                new_cost = float(row['COGS'])
            except (ValueError, TypeError):
                continue

            if not existing_entry.empty:
                old_cost = existing_entry.iloc[0]['Cost']
                sku = existing_entry.iloc[0]['SKU']
                
                if pd.isna(old_cost) or old_cost == '':
                    # Update empty cost and record actual update
                    sb_df.loc[sb_df['ASIN'] == asin, 'Cost'] = new_cost
                    actual_updates.append({
                        'ASIN': asin,
                        'SKU': sku,
                        'Name': name,
                        'new_cost': new_cost
                    })
                else:
                    # Record potential update
                    potential_updates.append({
                        'ASIN': asin,
                        'SKU': sku,
                        'Name': name,
                        'old_cost': float(old_cost),
                        'new_cost': new_cost
                    })
            else:
                # Add new product
                sku = generate_sku()
                new_products.append({
                    'ASIN': asin,
                    'SKU': sku,
                    'Name': name,
                    'cost': new_cost
                })
                
                # Add to SB DataFrame (only new products)
                new_sb_row = {
                    'ASIN': asin,
                    'SKU': sku,
                    'Title': name,
                    'Labels': '#FBA',
                    'Cost': new_cost,
                    'VAT_CATEGORY': 'A_GEN_STANDARD',
                    'Hide': 'NO'
                }
                sb_df = pd.concat([sb_df, pd.DataFrame([new_sb_row])], ignore_index=True)
                
                # Add to Listing Loader (only new products)
                new_row = [None] * len(headers)
                new_row[col_indices['Your Search Term'] - 1] = asin
                new_row[col_indices['Recommended Action'] - 1] = 'Ready To list > Enter required details.'
                new_row[col_indices["Amazon's Title"] - 1] = row['Name']
                new_row[col_indices['Record Action'] - 1] = 'Add Product'
                new_row[col_indices['Seller SKU'] - 1] = sku
                new_row[col_indices['Merchant Suggested ASIN'] - 1] = asin
                new_row[col_indices['Offering Condition Type'] - 1] = 'New'
                new_row[col_indices['Fulfillment Channel Code (US)'] - 1] = 'AMAZON_NA'
                new_row[col_indices['Your Price USD (Sell on Amazon, US)'] - 1] = f"{float(row['Sale Price']) * 1.15:.2f}"
                ws.append(new_row)

        listing_loader_output_buffer = io.BytesIO()  # Create a memory buffer
        wb.save(listing_loader_output_buffer)  # Save to the buffer
        listing_loader_output_buffer.seek(0)  # Rewind to the beginning of the buffer

        # Create a BytesIO buffer for the updated SB file in .xlsx format
        sb_buffer = io.BytesIO()
        sb_df.to_excel(sb_buffer, index=False, engine='openpyxl')  # Save in .xlsx format
        sb_buffer.seek(0)  # Rewind to the beginning of the buffer

        # Upload updated SB file to S3
        s3_client = boto3.client('s3')
        try:
            s3_client.put_object(
                Bucket=CONFIG_S3_BUCKET,
                Key=SB_FILE_KEY,
                Body=sb_buffer.getvalue()
            )
            print("Successfully uploaded updated SB file to S3")
        except Exception as e:
            print(f"Error uploading SB file to S3: {e}")
            raise

        # Reset buffer position for email attachment
        sb_buffer.seek(0)

        # Send email with both files as attachments
        send_email(BytesIO(listing_loader_output_buffer.read()), 
                  "listingLoaderUpdated.xlsm", 
                  sb_buffer, SB_UPDATED_FILE,
                  "tevinz159@gmail.com",
                  potential_updates, new_products, actual_updates)

        # Update last processed date after the process
        new_last_processed_date = str(leads_df["Date"].max().date())
        update_last_processed_date(new_last_processed_date)

        return {
            'statusCode': 200,
            'body': json.dumps('Process completed successfully!')
        }
    
    except Exception as e:
        print(f"Error: {e}")
        return {
            'statusCode': 500,
            'body': json.dumps(f"Error: {str(e)}")
        }